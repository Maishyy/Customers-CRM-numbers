# streamlit_app.py
# Updated with:
# 1. Multi-file upload support in Daily SMS tab
# 2. Improved business logic for customer messaging eligibility
# 3. 14-day cooldown period for existing customers
# 4. Automatic new customer registration

import os
import re
import json
import logging
from io import BytesIO
from textwrap import dedent
from datetime import datetime, timedelta
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache

import pandas as pd
import streamlit as st
import pdfplumber

import firebase_admin
from firebase_admin import credentials, firestore

# -----------------------------
# Streamlit page config
# -----------------------------
st.set_page_config(layout="wide", page_title="Sequid Hardware Contact System")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("app.log")]
)
logger = logging.getLogger(__name__)

# -----------------------------
# Firebase Setup
# -----------------------------
TEST_MODE = st.sidebar.checkbox(
    "Enable Test Mode",
    help="Use local Firebase emulator (set FIRESTORE_EMULATOR_HOST)."
)

db = None
if TEST_MODE:
    os.environ["FIRESTORE_EMULATOR_HOST"] = "localhost:8080"
    try:
        if not firebase_admin._apps:
            if os.path.exists("local_test_creds.json"):
                firebase_admin.initialize_app(credentials.Certificate("local_test_creds.json"))
            else:
                firebase_admin.initialize_app()
        db = firestore.client()
        st.sidebar.warning("TEST MODE ACTIVE - Using local emulator")
    except Exception as e:
        logger.error(f"Firebase (emulator) initialization failed: {str(e)}")
        st.error("Failed to initialize Firebase emulator. Check logs for details.")
        st.stop()
else:
    try:
        firebase_json = json.loads(st.secrets["firebase_creds"])
        cred = credentials.Certificate(firebase_json)
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        db = firestore.client()
    except Exception as e:
        logger.error(f"Firebase initialization failed: {str(e)}")
        st.error("Failed to initialize Firebase (check st.secrets['firebase_creds']).")
        st.stop()

# -----------------------------
# Constants & Regex
# -----------------------------

# Flexible phone matcher (accepts +254 7xx xxx xxx, 07xxxxxxxx, 7xxxxxxxx, 011xxxxxxxx; spaces/dashes allowed)
PHONE_FLEX = re.compile(
    r'(?:\+?254|0)?\s*(?:7\d{2}|1[0-9]\d)\s*[\s-]?\d{3}[\s-]?\d{3}'
)

# Bank-specific relaxed matcher (same idea, kept separate for clarity)
BANK_PHONE_REGEX = re.compile(
    r'(?:\+?254|0)?\s*(?:7\d{2}|1[0-9]\d)\s*\d{3}\s*\d{3}'
)

# Name patterns: allow 1–3 tokens, letters plus apostrophes/hyphens, mixed case
NAME_TOKEN = r"[A-Za-z][A-Za-z\'\-]{1,}"
NAME_PATTERN = re.compile(rf'({NAME_TOKEN}(?:\s+{NAME_TOKEN}){{0,2}})')

# Noise & codes to strip
NOISE_TOKEN = re.compile(r'\b(?:MPESA|M-PESA|PAYBILL|TILL|ACC(?:OUNT)?|REF(?:ERENCE)?|BALANCE|CONFIRMATION|RECEIPT|SALES)\b', re.I)
TXN_CODE = re.compile(r'\b[A-Z0-9]{8,12}\b')  # generic MPESA/Bank-like refs (e.g., TGU53XDSZD)

# Typical MPESA refs start with letters (T...), include alphanumerics; we'll search last code and take name after it
MPESA_CODE_CHUNK = re.compile(r'\bT[A-Z0-9]{8,}\b', re.I)

EXCLUDE_WORDS_HARD = {"CDM", "CHEQUE"}
EXCLUDE_WORDS_SOFT = {"BANK", "TRANSFER", "LTD", "LIMITED", "INTERNATIONAL"}

BLACKLISTED_NUMBERS = {"+254722000000", "+254000000000", "0766145780", "+254722000000", "254722000000", "0722000000"}
MAX_FILE_SIZE_MB = 10

# Kenyan mobile prefixes: 07x and 01x (0110–0119, etc.). We'll validate more generically.
def _valid_ke_prefix(digits12: str) -> bool:
    # digits12: '2547xxxxxxxx' or '2541xxxxxxxx'
    if len(digits12) != 12 or not digits12.startswith('254'):
        return False
    after = digits12[3:]  # e.g., '7xxxxxxxxx' or '1xxxxxxxxx'
    if after[0] == '7' and after[1].isdigit():
        return True
    if after[0] == '1' and after[1].isdigit():
        return True
    return False

# -----------------------------
# Helpers
# -----------------------------
def safe_file_size(uploaded_file) -> int:
    """Return file size in bytes in a Streamlit-version-safe way."""
    try:
        return uploaded_file.size  # may exist on newer versions
    except Exception:
        try:
            return len(uploaded_file.getbuffer())
        except Exception:
            try:
                pos = uploaded_file.tell()
            except Exception:
                pass
            try:
                uploaded_file.seek(0, 2)
                size = uploaded_file.tell()
                uploaded_file.seek(0)
                return size
            except Exception:
                return 0

def format_phone_number(raw: str):
    """Standardize phone to +254 format with rigorous validation."""
    if not raw:
        return None
    digits = re.sub(r'\D', '', str(raw))

    # Normalize to 254xxxxxxxxx (12 digits with country code)
    if digits.startswith('0') and len(digits) == 10:
        digits = '254' + digits[1:]
    elif digits.startswith('7') and len(digits) == 9:
        digits = '254' + digits
    elif digits.startswith('1') and len(digits) == 9:
        digits = '254' + digits
    elif digits.startswith('254') and len(digits) == 12:
        pass
    else:
        return None

    if not _valid_ke_prefix(digits):
        return None

    phone = '+' + digits
    if phone in BLACKLISTED_NUMBERS:
        return None
    return phone

def clean_name(name: str) -> str:
    """Clean and normalize human-like names."""
    if not name or not isinstance(name, str):
        return ""

    # Remove obvious noise and refs
    name = NOISE_TOKEN.sub(' ', name)
    name = TXN_CODE.sub(' ', name)
    # Remove non-letters except space, apostrophe, hyphen
    name = re.sub(r"[^A-Za-z'\-\s]", ' ', name)
    # Collapse spaces
    name = re.sub(r'\s+', ' ', name).strip()

    tokens = [t.capitalize() for t in name.split() if t]
    # de-duplicate tokens preserving order
    seen = set()
    uniq = []
    for t in tokens:
        tl = t.lower()
        if tl not in seen:
            seen.add(tl)
            uniq.append(t)

    # Keep up to 3 tokens
    uniq = uniq[:3]
    cleaned = ' '.join(uniq).strip()

    if not cleaned:
        return ""
    if any(len(tok) < 2 for tok in cleaned.split()):
        return ""
    return cleaned

def should_exclude_line(line: str, has_phone: bool) -> bool:
    """Exclude headers or noise lines. Be softer if the line has a phone."""
    if not isinstance(line, str):
        return True
    U = line.upper()
    if any(w in U for w in EXCLUDE_WORDS_HARD):
        return True
    if not has_phone and any(w in U for w in EXCLUDE_WORDS_SOFT):
        return True
    return False

def try_extract_name_near(text_line: str, start_idx: int, end_idx: int) -> str:
    """Grab a plausible name near the phone in the same line (left or right window)."""
    window_left = text_line[max(0, start_idx - 50):start_idx]
    window_right = text_line[end_idx:end_idx + 50]

    m_left = NAME_PATTERN.findall(window_left)
    cand = m_left[-1] if m_left else ''
    if not cand:
        m_right = NAME_PATTERN.findall(window_right)
        cand = m_right[0] if m_right else ''
    return clean_name(cand)

def validate_contact(phone, name):
    if not phone or len(phone) != 13 or not phone.startswith("+254"):
        return False
    digits = phone[1:]
    if not _valid_ke_prefix(digits):
        return False
    if name:
        parts = name.split()
        if any(len(p) < 2 for p in parts):
            return False
    return True

def validate_contact_strict(phone, name):
    if not validate_contact(phone, name):
        return False
    digits = re.sub(r'\D', '', phone)
    if len(set(digits)) < 4:  # prevent things like +254777777777
        return False
    return True

# -----------------------------
# Extraction (generic)
# -----------------------------
@lru_cache(maxsize=64)
def extract_contacts(text: str):
    """Extract contacts line-by-line with improved heuristics and validation."""
    results = []
    seen = set()
    lines = re.split(r'\r?\n+', text)

    for idx, line in enumerate(lines):
        if not isinstance(line, str) or not line.strip():
            continue

        phones = list(PHONE_FLEX.finditer(line))
        has_phone = bool(phones)
        if should_exclude_line(line, has_phone):
            continue

        # Skip rows with many random numbers (likely refs)
        if len(re.findall(r'\d+', line)) > 6 and not has_phone:
            continue

        for m in phones:
            raw = m.group(0)
            phone = format_phone_number(raw)
            if not phone or phone in seen:
                continue

            name = try_extract_name_near(line, m.start(), m.end())

            # Fallback to previous line
            if not name and idx > 0:
                nm_prev = NAME_PATTERN.search(lines[idx - 1]) if idx > 0 else None
                if nm_prev:
                    name = clean_name(nm_prev.group(0))

            if name and not validate_contact(phone, name):
                name = ""

            if validate_contact(phone, name or ""):
                results.append((phone, name or ""))
                seen.add(phone)

    return results

# -----------------------------
# Bank-specific extraction (Co-op & Equity)
# -----------------------------
def _guess_narrative_columns(df: pd.DataFrame):
    # Prefer explicit "Narrative" (case-insensitive)
    narrative_cols = [c for c in df.columns if isinstance(c, str) and 'narrative' in c.lower()]
    if narrative_cols:
        return narrative_cols

    # Heuristic: pick text-like columns with long strings and MPESA-like patterns
    candidates = []
    for c in df.columns:
        try:
            series = df[c].dropna().astype(str)
        except Exception:
            continue
        if series.empty:
            continue
        # check typical patterns (MPS, Txxxx codes, 2547…)
        sample = " ".join(series.head(10).tolist())
        hits = 0
        if re.search(r'\bMPS\b|\bM-PESA\b|\bMPESA\b', sample, re.I):
            hits += 1
        if re.search(r'\bT[A-Z0-9]{8,}\b', sample, re.I):
            hits += 1
        if re.search(r'\b2547\d{7}\b', sample):
            hits += 1
        if hits >= 1:
            candidates.append(c)

    # If still nothing, pick the longest average-length text column
    if not candidates:
        best_col = None
        best_len = 0
        for c in df.columns:
            try:
                series = df[c].dropna().astype(str)
            except Exception:
                continue
            if series.empty:
                continue
            avg_len = series.map(len).mean()
            if avg_len > best_len:
                best_len = avg_len
                best_col = c
        if best_col:
            return [best_col]
        return []

    return candidates

def extract_from_bank_statement(df: pd.DataFrame):
    """Special parser for bank statement Excel with a Narrative-like column."""
    results = []
    seen = set()

    narrative_cols = _guess_narrative_columns(df)
    if not narrative_cols:
        return results

    for _, row in df.iterrows():
        for col in narrative_cols:
            val = str(row.get(col, "")) if pd.notna(row.get(col, "")) else ""
            if not val.strip():
                continue

            # First try the improved pattern for ~ separated format
            improved_pattern = re.compile(r'(?:~|\b)(254[17]\d{8})(?:~|\b).*?(?:~|\b)([A-Za-z][A-Za-z\s]+)(?=\s*~|\b|$)')
            improved_matches = improved_pattern.findall(val)
            
            for phone, name in improved_matches:
                formatted_phone = f"+{phone}"
                if formatted_phone in seen:
                    continue
                cleaned_name = clean_name(name)
                if validate_contact(formatted_phone, cleaned_name or ""):
                    results.append((formatted_phone, cleaned_name))
                    seen.add(formatted_phone)
                continue  # Skip other processing if we found a match

            # Fall back to original processing if improved pattern didn't match
            phones = BANK_PHONE_REGEX.findall(val)
            phones = [format_phone_number(p) for p in phones]
            phones = [p for p in phones if p]

            if not phones:
                continue

            phone = phones[0]
            if phone in seen:
                continue

            # Handle the specific case where name comes after phone and MPESA code
            name_part = val
            if '~' in val:
                parts = val.split('~')
                for i, part in enumerate(parts):
                    if phone[1:] in part:  # phone without '+'
                        # Look ahead for name in next parts
                        for j in range(i+1, len(parts)):
                            if re.match(r'^[A-Za-z]', parts[j]):
                                name_part = parts[j]
                                break
                        break

            name = clean_name(name_part)

            # If no decent name found, try alternative patterns
            if not name:
                m = re.search(rf'{re.escape(phone[1:])}[^A-Za-z]*([A-Za-z][A-Za-z\s]+)', val)
                if m:
                    name = clean_name(m.group(1))

            if validate_contact(phone, name or ""):
                results.append((phone, name))
                seen.add(phone)

    return results

# -----------------------------
# File processors
# -----------------------------
def extract_from_dataframe(df: pd.DataFrame):
    records = []
    for _, row in df.iterrows():
        vals = [str(x) for x in row.values.tolist() if pd.notna(x)]
        line = " ".join(vals)
        has_phone = bool(PHONE_FLEX.search(line))
        if not should_exclude_line(line, has_phone):
            records.extend(extract_contacts(line))
    return records

def _read_excel_all_sheets(file):
    """Read all sheets with correct engine based on extension, with graceful errors."""
    name = (getattr(file, "name", "") or "").lower()
    engine = None
    if name.endswith(".xlsx"):
        engine = "openpyxl"
    elif name.endswith(".xls"):
        # For .xls, pandas needs xlrd installed. Let pandas attempt default engine first,
        # and if it errors about xlrd, we surface a helpful message.
        engine = None

    try:
        return pd.read_excel(file, sheet_name=None, engine=engine)
    except Exception as e:
        msg = str(e)
        if ".xls" in name and "xlrd" in msg.lower():
            st.error(
                "This .xls file requires the optional dependency **xlrd**. "
                "Please install it in your environment, e.g.:\n\n"
                "`pip install xlrd`"
            )
        raise

def process_pdf(file):
    try:
        with pdfplumber.open(file) as pdf:
            total_pages = len(pdf.pages)
            progress_bar = st.progress(0)
            contacts = []
            seen_numbers = set()

            # Basic sampling for very long PDFs
            pages = pdf.pages[::2] if total_pages > 20 else pdf.pages

            for i, page in enumerate(pages):
                try:
                    txt = page.extract_text() or ""
                except Exception:
                    txt = ""
                if txt:
                    for phone, name in extract_contacts(txt):
                        if phone and phone not in seen_numbers:
                            contacts.append((phone, name))
                            seen_numbers.add(phone)
                progress_bar.progress((i + 1) / len(pages))
            return contacts
    except Exception as e:
        logger.error(f"PDF processing error: {str(e)}")
        st.error(f"Error processing PDF: {str(e)}")
        return []

def process_csv(file):
    try:
        size = safe_file_size(file)
        if size > 5 * 1024 * 1024:  # 5MB
            records = []
            for chunk in pd.read_csv(file, chunksize=10000):
                records.extend(extract_from_dataframe(chunk))
            return records
        return extract_from_dataframe(pd.read_csv(file))
    except Exception as e:
        logger.error(f"CSV processing error: {str(e)}")
        st.error(f"Error processing CSV: {str(e)}")
        return []

def process_excel(file):
    try:
        dfs = _read_excel_all_sheets(file)
        output = []
        for sheet_name, df in dfs.items():
            cols_lower = [str(c).lower() for c in df.columns]
            # If sheet has/looks like "Narrative", apply bank-specific parsing
            if any('narrative' in c for c in cols_lower) or _guess_narrative_columns(df):
                output.extend(extract_from_bank_statement(df))
            else:
                output.extend(extract_from_dataframe(df))
        return output
    except Exception as e:
        logger.error(f"Excel processing error: {str(e)}")
        st.error(f"Error processing Excel: {str(e)}")
        return []

def process_file(file):
    try:
        ext = (file.name or "").lower().split('.')[-1]
        if ext == "pdf":
            return process_pdf(file)
        elif ext == "csv":
            return process_csv(file)
        elif ext in ("xls", "xlsx"):
            return process_excel(file)
        else:
            st.warning(f"Unsupported file format: {file.name}")
            return []
    except Exception as e:
        logger.error(f"Error processing {file.name}: {str(e)}")
        st.error(f"Error processing {file.name}: {str(e)}")
        return []

def process_files_parallel(files):
    with ThreadPoolExecutor() as executor:
        results = list(executor.map(process_file, files))
    return [c for sub in results for c in sub]

def process_file_with_duplicate_checks(file):
    try:
        contacts = process_file(file)
        unique_contacts = {}
        for phone, name in contacts:
            if not phone:
                continue
            k = phone.replace("+", "").strip()
            nm = clean_name(name)
            if k in unique_contacts:
                # prefer longer name
                if len(nm) > len(unique_contacts[k][1]):
                    unique_contacts[k] = (phone, nm)
            else:
                unique_contacts[k] = (phone, nm)
        return list(unique_contacts.values())
    except Exception as e:
        logger.error(f"Error in duplicate check: {str(e)}")
        return []

# -----------------------------
# Firebase Operations
# -----------------------------
def save_to_firestore(data):
    """Save contacts to Firestore with duplicate prevention."""
    if not db or not data:
        return 0, 0

    coll = db.collection("contacts")
    batch = db.batch()
    new_count = 0
    duplicate_count = 0

    # Preload existing numbers
    existing_numbers = set()
    try:
        for doc in coll.select(["phone_number"]).stream():
            existing_numbers.add(doc.get("phone_number"))
    except Exception as e:
        logger.warning(f"Prefetch existing numbers failed: {e}")

    for phone, name in {(p, n) for p, n in data if p}:
        if not validate_contact_strict(phone, name):
            continue

        if phone in existing_numbers:
            duplicate_count += 1
            continue

        doc_ref = coll.document(phone.replace("+", ""))

        first, last = "", ""
        if name:
            parts = name.strip().split(" ", 1)
            first = parts[0]
            last = parts[1] if len(parts) > 1 else ""

        batch.set(doc_ref, {
            "phone_number": phone,
            "client_name": name,
            "first_name": first,
            "last_name": last,
            "source": "upload",
            "timestamp": firestore.SERVER_TIMESTAMP,
            "last_transaction_date": firestore.SERVER_TIMESTAMP
        })
        new_count += 1
        existing_numbers.add(phone)

        if new_count % 500 == 0:
            batch.commit()
            batch = db.batch()

    if new_count % 500 != 0:
        batch.commit()

    return new_count, duplicate_count

def log_message(phone, name):
    try:
        db.collection("messages_sent").add({
            "phone_number": phone,
            "client_name": name,
            "timestamp": firestore.SERVER_TIMESTAMP,
            "status": "queued"
        })
    except Exception as e:
        logger.error(f"Failed to log message for {phone}: {str(e)}")

@st.cache_data(ttl=3600)
def load_message_logs(days=30):
    try:
        cutoff = datetime.now() - timedelta(days=days)
        docs = (db.collection("messages_sent")
                .where("timestamp", ">=", cutoff)
                .stream())
        data = []
        for doc in docs:
            d = doc.to_dict()
            phone = d.get("phone_number", "")
            ts = d.get("timestamp")
            # normalize timestamp display
            if hasattr(ts, 'strftime'):
                date_str = ts.strftime("%Y-%m-%d")
            else:
                try:
                    date_str = ts.to_datetime().strftime("%Y-%m-%d") if ts else ""
                except Exception:
                    date_str = ""
            data.append({
                "Phone": phone,
                "Date Messaged": date_str,
                "Status": d.get("status", "unknown")
            })
        return pd.DataFrame(data)
    except Exception as e:
        logger.error(f"Error loading logs: {str(e)}")
        st.error("Could not load message logs.")
        return pd.DataFrame(columns=["Phone", "Date Messaged", "Status"])

def get_last_message_dates(phone_numbers):
    last_messages = {}
    if not phone_numbers:
        return last_messages

    batch_size = 30
    batches = [phone_numbers[i:i + batch_size] for i in range(0, len(phone_numbers), batch_size)]
    for batch in batches:
        if not batch:
            continue
        try:
            for p in batch:
                docs = (db.collection("messages_sent")
                        .where("phone_number", "==", p)
                        .order_by("timestamp", direction=firestore.Query.DESCENDING)
                        .limit(1)
                        .stream())
                for doc in docs:
                    data = doc.to_dict()
                    last_messages[data["phone_number"]] = data.get("timestamp")
        except Exception as e:
            logger.error(f"Error fetching messages for batch: {str(e)}")
            continue
    return last_messages

# -----------------------------
# Export / Download
# -----------------------------
def generate_standard_excel(data):
    """Generate standardized Excel output with openpyxl styling."""
    formatted_data = []
    seen = set()
    for phone, name in data:
        if phone in seen:
            continue
        seen.add(phone)
        first = last = ""
        if name:
            parts = name.strip().split()
            first = parts[0] if parts else ""
            last = " ".join(parts[1:]) if len(parts) > 1 else ""
        clean_phone = phone.replace("+", "") if phone else ""
        formatted_data.append({
            "Firstname(optional)": first,
            "Lastname(optional)": last,
            "Phone or Email": clean_phone,
            "phone(254)": clean_phone,
            "Valid": "Yes" if validate_contact(phone, name) else "No"
        })

    df = pd.DataFrame(formatted_data)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Contacts')
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = writer.book
        ws = writer.sheets['Contacts']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_alignment = Alignment(vertical="top", wrap_text=True)
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header styling
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Autosize columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()])
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buffer.seek(0)
    return buffer, df

def download_full_contact_list():
    try:
        contacts_ref = db.collection("contacts")
        docs = contacts_ref.stream()
        contact_list = []
        for doc in docs:
            data = doc.to_dict()
            contact_list.append((data.get("phone_number", ""), data.get("client_name", "")))
        return generate_standard_excel(contact_list)
    except Exception as e:
        logger.error(f"Error downloading full contact list: {str(e)}")
        return None, None

# -----------------------------
# UI
# -----------------------------
tabs = st.tabs(["Upload & Sync", "Daily SMS List", "Dashboard", "Data Quality"])

# --- Tab 1: Upload & Sync ---
with tabs[0]:
    st.subheader("Upload MPESA or Bank Statements")
    uploaded_files = st.file_uploader(
        "Upload files",
        type=["pdf", "csv", "xls", "xlsx"],
        accept_multiple_files=True,
        help=f"Max file size: {MAX_FILE_SIZE_MB}MB each"
    )

    if uploaded_files:
        oversized = [f.name for f in uploaded_files if safe_file_size(f) > MAX_FILE_SIZE_MB * 1024 * 1024]
        if oversized:
            st.error(f"Oversized files: {', '.join(oversized)}")
        else:
            with st.spinner("Processing files..."):
                all_data = process_files_parallel(uploaded_files) if len(uploaded_files) > 1 else process_file(uploaded_files[0])

            if all_data:
                preview_df = pd.DataFrame(all_data, columns=["Phone", "Name"])
                preview_df["Valid"] = preview_df.apply(
                    lambda row: "✅" if validate_contact(row["Phone"], row["Name"]) else "❌",
                    axis=1
                )
                st.subheader("Data Quality Preview")
                st.dataframe(preview_df.head(50))

                unique_count = len({p for p, _ in all_data if p})
                st.success(f"Found {unique_count} unique contacts.")

                if st.button("Confirm Upload to Database"):
                    with st.spinner("Saving to database..."):
                        new_count, duplicate_count = save_to_firestore(all_data)
                    st.success(f"Added {new_count} new contacts; skipped {duplicate_count} duplicates.")

                    excel_file, df = generate_standard_excel(all_data)
                    st.download_button(
                        "Download Processed Contacts",
                        data=excel_file,
                        file_name="processed_contacts.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("No valid phone numbers found.")

    st.subheader("Contact Management")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Download Full Contact List"):
            with st.spinner("Preparing full contact list..."):
                excel_file, df = download_full_contact_list()
                if excel_file:
                    st.download_button(
                        label="Download Complete Contacts",
                        data=excel_file,
                        file_name="full_contact_list.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.dataframe(df.head())
                else:
                    st.error("Failed to generate contact list")

    with col2:
        if st.button("Refresh Contact Count"):
            try:
                agg = db.collection("contacts").count().get()
                if isinstance(agg, list):
                    count = agg[0][0].value
                else:
                    # try to get .value or nested aggregation result in some SDK variants
                    count = getattr(agg, "value", None) or getattr(getattr(agg, "aggregation_results", [{}])[0], "value", None)
                    if isinstance(count, dict) and "integerValue" in count:
                        count = int(count["integerValue"])
                if count is None:
                    raise ValueError("Count unavailable; falling back.")
            except Exception:
                count = sum(1 for _ in db.collection("contacts").stream())
            st.metric("Total Contacts in Database", count)

# --- Tab 2: Daily SMS List ---
with tabs[1]:
    st.subheader("Generate Daily SMS List")
    
    # Configurable parameters
    with st.expander("Messaging Settings"):
        COOLDOWN_DAYS = st.slider(
            "Minimum days between messages", 
            min_value=1, 
            max_value=30, 
            value=14,
            help="Customers won't receive messages more frequently than this"
        )
        INCLUDE_NEW = st.checkbox(
            "Always include new customers", 
            value=True,
            help="Automatically add phone numbers not in our database"
        )
    
    # Multi-file uploader
    uploaded_files = st.file_uploader(
        "Upload today's statements",
        type=["pdf", "csv", "xls", "xlsx"],
        accept_multiple_files=True,
        key="daily_sms_uploader"
    )

    if uploaded_files and st.button("Generate Today's SMS List"):
        with st.spinner("Processing statements..."):
            # Process all files and combine contacts
            all_contacts = []
            for file in uploaded_files:
                file_contacts = process_file_with_duplicate_checks(file)
                all_contacts.extend(file_contacts)
            
            # Remove duplicates from combined list
            unique_contacts = {p: n for p, n in all_contacts if p}
            st.info(f"Processed {len(uploaded_files)} files with {len(unique_contacts)} unique contacts")
            
            # Prepare Firestore operations
            sms_list = []
            batch = db.batch()
            contacts_ref = db.collection("contacts")
            now = datetime.now()
            
            # Process each contact
            for phone, name in unique_contacts.items():
                doc_ref = contacts_ref.document(phone.replace("+", ""))
                
                # Try to get existing record
                doc = doc_ref.get()
                
                eligible = False
                if doc.exists:
                    # Existing customer - check last transaction date
                    last_trans = doc.to_dict().get("last_transaction_date")
                    if isinstance(last_trans, datetime):
                        delta = now - last_trans
                        eligible = delta.days > COOLDOWN_DAYS
                    else:
                        eligible = True  # if field missing, include to be safe
                else:
                    # New customer - check if we should include
                    eligible = INCLUDE_NEW
                
                if eligible:
                    sms_list.append((phone, name))
                    
                    # Prepare update data
                    update_data = {
                        "last_transaction_date": firestore.SERVER_TIMESTAMP,
                        "phone_number": phone,
                        "client_name": name
                    }
                    
                    # Additional fields for new customers
                    if not doc.exists:
                        name_parts = name.split() if name else []
                        update_data.update({
                            "first_name": name_parts[0] if name_parts else "",
                            "last_name": " ".join(name_parts[1:]) if len(name_parts) > 1 else "",
                            "source": "upload",
                            "timestamp": firestore.SERVER_TIMESTAMP
                        })
                    
                    # Queue the update
                    batch.set(doc_ref, update_data, merge=True)
            
            # Commit all updates
            batch.commit()
            
            # Log messages for eligible contacts
            for phone, name in sms_list:
                log_message(phone, name)
            
            # Show results
            if sms_list:
                st.success(f"{len(sms_list)} contacts eligible for messaging (cooldown: {COOLDOWN_DAYS} days)")
                
                # Generate downloadable report
                sms_excel, sms_df = generate_standard_excel(sms_list)
                st.download_button(
                    "Download SMS List",
                    data=sms_excel,
                    file_name=f"sms_list_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Show preview
                st.dataframe(sms_df)
            else:
                st.info("No eligible contacts to message today from these statements.")

# --- Tab 3: Dashboard ---
with tabs[2]:
    st.subheader("Message History Dashboard")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start date", datetime.now() - timedelta(days=30))
    with col2:
        end_date = st.date_input("End date", datetime.now())

    if st.button("Refresh Dashboard"):
        df_logs = load_message_logs((end_date - start_date).days)
        if df_logs.empty:
            st.info("No message logs found for selected period.")
        else:
            total_messages = len(df_logs)
            unique_contacts = df_logs["Phone"].nunique()
            success_rate = len(df_logs[df_logs["Status"].str.lower() == "delivered"]) / total_messages if total_messages else 0

            c1, c2, c3 = st.columns(3)
            c1.metric("Total Messages", total_messages)
            c2.metric("Unique Contacts", unique_contacts)
            c3.metric("Success Rate", f"{success_rate:.1%}")

            st.subheader("Message Volume")
            daily_counts = df_logs.groupby("Date Messaged").size()
            st.bar_chart(daily_counts)

            st.subheader("Top Recipients")
            top_contacts = df_logs["Phone"].value_counts().nlargest(10)
            st.bar_chart(top_contacts)

            st.subheader("Message Log")
            st.dataframe(df_logs.sort_values("Date Messaged", ascending=False))

            csv = df_logs.to_csv(index=False).encode("utf-8")
            st.download_button(
                "Download Full Log CSV",
                data=csv,
                file_name="message_log.csv",
                mime="text/csv"
            )

# --- Tab 4: Data Quality ---
with tabs[3]:
    st.subheader("Data Quality Analysis")

    if st.button("Run Quality Check"):
        with st.spinner("Analyzing database..."):
            contacts_ref = db.collection("contacts")
            docs = contacts_ref.stream()

            stats = {
                "total_contacts": 0,
                "invalid_phones": 0,
                "missing_names": 0,
                "duplicate_phones": defaultdict(int),
                "phone_prefixes": defaultdict(int)
            }
            phone_counts = defaultdict(int)

            for doc in docs:
                data = doc.to_dict()
                phone = data.get("phone_number", "")
                name = data.get("client_name", "")

                stats["total_contacts"] += 1
                phone_counts[phone] += 1

                if not validate_contact(phone, name or ""):
                    stats["invalid_phones"] += 1
                if not (name or "").strip():
                    stats["missing_names"] += 1

                if phone.startswith("+254") and len(phone) >= 6:
                    pref2 = phone[4:6]  # shows '7x' or '1x'
                    stats["phone_prefixes"][pref2] += 1

            stats["duplicate_count"] = sum(1 for count in phone_counts.values() if count > 1)

            st.metric("Total Contacts", stats["total_contacts"])

            col1, col2 = st.columns(2)
            col1.metric("Invalid Phone Numbers", stats["invalid_phones"])
            col2.metric("Contacts Missing Names", stats["missing_names"])

            st.metric("Duplicate Phone Numbers", stats["duplicate_count"])

            st.subheader("Phone Prefix Distribution")
            prefix_df = pd.DataFrame.from_dict(
                stats["phone_prefixes"], orient="index", columns=["Count"]
            ).sort_values("Count", ascending=False)
            st.bar_chart(prefix_df)

            if stats["duplicate_count"] > 0:
                st.subheader("Sample Duplicate Entries")
                duplicates = {p: c for p, c in phone_counts.items() if c > 1}
                sample_dupes = list(duplicates.items())[:10]
                for phone, count in sample_dupes:
                    st.write(f"{phone}: {count} occurrences")
                    dlist = contacts_ref.where("phone_number", "==", phone).stream()
                    for dd in dlist:
                        st.write(f"- {dd.to_dict().get('client_name', 'No name')}")