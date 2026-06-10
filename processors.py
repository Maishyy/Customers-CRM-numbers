# processors.py
from io import BytesIO
import logging
import re
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache

import pandas as pd
import pdfplumber
import streamlit as st

from config import (
    PHONE_FLEX,
    BANK_PHONE_REGEX,
    NAME_PATTERN,
    EXCLUDE_WORDS_HARD,
    EXCLUDE_WORDS_SOFT,
    SMS_FAILED_STATUSES,
    SMS_NON_PROMOTIONAL_STATUSES,
    SMS_PROMOTIONAL_STATUSES,
)
from utils import format_phone_number, clean_name, validate_contact, safe_file_size

logger = logging.getLogger(__name__)

COOP_NARRATION_PATTERN = re.compile(
    r'~\s*(254[17]\d{8})\s*~[^~]*~\s*([A-Za-z][A-Za-z\'\-\s]+?)(?=\s*(?:~|$))',
    re.I,
)
FAMILY_REMARKS_PATTERN = re.compile(
    r'\bFrom\s+(254[17]\d{8})\s+([A-Za-z][A-Za-z\'\-\s]+?)(?=\s+Alias\s+Code\b|$)',
    re.I,
)

SMS_STATUS_LOOKUP = {
    status.lower().replace(" ", ""): status
    for status in SMS_PROMOTIONAL_STATUSES | SMS_NON_PROMOTIONAL_STATUSES | SMS_FAILED_STATUSES
}

def should_exclude_line(line: str, has_phone: bool) -> bool:
    """Exclude headers or noise lines."""
    if not isinstance(line, str):
        return True
    U = line.upper()
    if any(w in U for w in EXCLUDE_WORDS_HARD):
        return True
    if not has_phone and any(w in U for w in EXCLUDE_WORDS_SOFT):
        return True
    return False

def try_extract_name_near(text_line: str, start_idx: int, end_idx: int) -> str:
    """Grab a plausible name near the phone in the same line."""
    window_left = text_line[max(0, start_idx - 50):start_idx]
    window_right = text_line[end_idx:end_idx + 50]

    m_left = NAME_PATTERN.findall(window_left)
    cand = m_left[-1] if m_left else ''
    if not cand:
        m_right = NAME_PATTERN.findall(window_right)
        cand = m_right[0] if m_right else ''
    return clean_name(cand)

@lru_cache(maxsize=64)
def extract_contacts(text: str):
    """Extract contacts line-by-line with improved heuristics and validation."""
    results = []
    seen = set()
    lines = re.split(r'\r?\n+', text)

    for idx, line in enumerate(lines):
        if not isinstance(line, str) or not line.strip():
            continue

        phones = list(PHONE_FLEX.finditer(line))
        has_phone = bool(phones)
        if should_exclude_line(line, has_phone):
            continue

        if len(re.findall(r'\d+', line)) > 6 and not has_phone:
            continue

        for m in phones:
            raw = m.group(0)
            phone = format_phone_number(raw)
            if not phone or phone in seen:
                continue

            name = try_extract_name_near(line, m.start(), m.end())

            if not name and idx > 0:
                nm_prev = NAME_PATTERN.search(lines[idx - 1]) if idx > 0 else None
                if nm_prev:
                    name = clean_name(nm_prev.group(0))

            if name and not validate_contact(phone, name):
                name = ""

            if validate_contact(phone, name or ""):
                results.append((phone, name or ""))
                seen.add(phone)

    return results

def _guess_narrative_columns(df: pd.DataFrame):
    narrative_keywords = (
        "narrative",
        "description",
        "details",
        "remarks",
        "remark",
        "particulars",
        "transaction details",
        "transaction_detail",
    )
    narrative_cols = [
        c for c in df.columns
        if isinstance(c, str) and any(keyword in c.lower() for keyword in narrative_keywords)
    ]
    if narrative_cols:
        return narrative_cols

    candidates = []
    for c in df.columns:
        try:
            series = df[c].dropna().astype(str)
        except Exception:
            continue
        if series.empty:
            continue
        sample = " ".join(series.head(10).tolist())
        hits = 0
        if re.search(r'\bMPS\b|\bM-PESA\b|\bMPESA\b', sample, re.I):
            hits += 1
        if re.search(r'\bT[A-Z0-9]{8,}\b', sample, re.I):
            hits += 1
        if re.search(r'\b254[17]\d{8}\b', sample):
            hits += 1
        if hits >= 1:
            candidates.append(c)

    if not candidates:
        best_col = None
        best_len = 0
        for c in df.columns:
            try:
                series = df[c].dropna().astype(str)
            except Exception:
                continue
            if series.empty:
                continue
            avg_len = series.map(len).mean()
            if avg_len > best_len:
                best_len = avg_len
                best_col = c
        if best_col:
            return [best_col]
        return []

    return candidates

def extract_from_bank_statement(df: pd.DataFrame):
    """Special parser for bank statement Excel with a Narrative-like column."""
    results = []
    seen = set()

    narrative_cols = _guess_narrative_columns(df)
    if not narrative_cols:
        return results

    for _, row in df.iterrows():
        for col in narrative_cols:
            val = str(row.get(col, "")) if pd.notna(row.get(col, "")) else ""
            if not val.strip():
                continue

            improved_pattern = re.compile(r'(?:~|\b)(254[17]\d{8})(?:~|\b).*?(?:~|\b)([A-Za-z][A-Za-z\s]+)(?=\s*~|\b|$)')
            improved_matches = improved_pattern.findall(val)

            for phone, name in improved_matches:
                formatted_phone = f"+{phone}"
                if formatted_phone in seen:
                    continue
                cleaned_name = clean_name(name)
                if validate_contact(formatted_phone, cleaned_name or ""):
                    results.append((formatted_phone, cleaned_name))
                    seen.add(formatted_phone)
            if improved_matches:
                continue

            phones = BANK_PHONE_REGEX.findall(val)
            phones = [format_phone_number(p) for p in phones]
            phones = [p for p in phones if p]

            if not phones:
                continue

            phone = phones[0]
            if phone in seen:
                continue

            name_part = val
            if '~' in val:
                parts = val.split('~')
                for i, part in enumerate(parts):
                    if phone[1:] in part:
                        for j in range(i+1, len(parts)):
                            if re.match(r'^[A-Za-z]', parts[j]):
                                name_part = parts[j]
                                break
                        break

            name = clean_name(name_part)

            if not name:
                m = re.search(rf'{re.escape(phone[1:])}[^A-Za-z]*([A-Za-z][A-Za-z\s]+)', val)
                if m:
                    name = clean_name(m.group(1))
            if not name:
                tail_after_phone = re.split(re.escape(phone[1:]), val, maxsplit=1)
                if len(tail_after_phone) > 1:
                    name = clean_name(tail_after_phone[1])

            if validate_contact(phone, name or ""):
                results.append((phone, name))
                seen.add(phone)

    return results

def extract_structured_statement_contacts(df: pd.DataFrame):
    """Scan likely text-bearing Excel cells for known statement formats."""
    results = []
    seen = set()

    for _, row in df.iterrows():
        for value in row.values.tolist():
            if pd.isna(value):
                continue
            text = str(value).strip()
            if not text:
                continue

            matches = []
            matches.extend(COOP_NARRATION_PATTERN.findall(text))
            matches.extend(FAMILY_REMARKS_PATTERN.findall(text))

            if not matches:
                continue

            for phone_digits, raw_name in matches:
                phone = format_phone_number(phone_digits)
                name = clean_name(raw_name)
                if not phone or phone in seen:
                    continue
                if validate_contact(phone, name or ""):
                    results.append((phone, name))
                    seen.add(phone)

    return results

def extract_from_dataframe(df: pd.DataFrame):
    records = []
    for _, row in df.iterrows():
        vals = [str(x) for x in row.values.tolist() if pd.notna(x)]
        line = " ".join(vals)
        has_phone = bool(PHONE_FLEX.search(line))
        if not should_exclude_line(line, has_phone):
            records.extend(extract_contacts(line))
    return records

def _read_excel_all_sheets(file):
    name = (getattr(file, "name", "") or "").lower()
    engine = None
    if name.endswith(".xlsx"):
        engine = "openpyxl"
    elif name.endswith(".xls"):
        engine = None

    try:
        return pd.read_excel(file, sheet_name=None, engine=engine)
    except Exception as e:
        msg = str(e)
        if ".xls" in name and "xlrd" in msg.lower():
            st.error(
                "This .xls file requires the optional dependency **xlrd**. "
                "Please install it in your environment, e.g.:\n\n"
                "`pip install xlrd`"
            )
        raise

def process_pdf(file):
    try:
        with pdfplumber.open(file) as pdf:
            total_pages = len(pdf.pages)
            progress_bar = st.progress(0)
            contacts = []
            seen_numbers = set()
            extracted_text_chars = 0

            pages = pdf.pages

            for i, page in enumerate(pages):
                try:
                    txt = page.extract_text() or ""
                except Exception:
                    txt = ""
                if txt:
                    extracted_text_chars += len(txt.strip())
                    for phone, name in extract_contacts(txt):
                        if phone and phone not in seen_numbers:
                            contacts.append((phone, name))
                            seen_numbers.add(phone)
                progress_bar.progress((i + 1) / len(pages))
            if not contacts and extracted_text_chars == 0:
                st.warning(
                    "This PDF appears to contain little or no extractable text. "
                    "It may be a scanned/image-based statement and could require OCR."
                )
            return contacts
    except Exception as e:
        logger.error(f"PDF processing error: {str(e)}")
        st.error(f"Error processing PDF: {str(e)}")
        return []

def process_csv(file):
    try:
        size = safe_file_size(file)
        if size > 5 * 1024 * 1024:
            records = []
            for chunk in pd.read_csv(file, chunksize=10000):
                records.extend(extract_from_dataframe(chunk))
            return records
        return extract_from_dataframe(pd.read_csv(file))
    except Exception as e:
        logger.error(f"CSV processing error: {str(e)}")
        st.error(f"Error processing CSV: {str(e)}")
        return []

def process_excel(file):
    try:
        dfs = _read_excel_all_sheets(file)
        output = []
        for sheet_name, df in dfs.items():
            cols_lower = [str(c).lower() for c in df.columns]
            structured_hits = extract_structured_statement_contacts(df)
            if structured_hits:
                output.extend(structured_hits)
            if any('narrative' in c for c in cols_lower) or _guess_narrative_columns(df):
                output.extend(extract_from_bank_statement(df))
            else:
                output.extend(extract_from_dataframe(df))
        return output
    except Exception as e:
        logger.error(f"Excel processing error: {str(e)}")
        st.error(f"Error processing Excel: {str(e)}")
        return []

def process_file(file):
    try:
        ext = (file.name or "").lower().split('.')[-1]
        if ext == "pdf":
            return process_pdf(file)
        elif ext == "csv":
            return process_csv(file)
        elif ext in ("xls", "xlsx"):
            return process_excel(file)
        else:
            st.warning(f"Unsupported file format: {file.name}")
            return []
    except Exception as e:
        logger.error(f"Error processing {file.name}: {str(e)}")
        st.error(f"Error processing {file.name}: {str(e)}")
        return []

def process_files_parallel(files):
    with ThreadPoolExecutor() as executor:
        results = list(executor.map(process_file, files))
    return [c for sub in results for c in sub]

def process_file_with_duplicate_checks(file):
    try:
        contacts = process_file(file)
        unique_contacts = {}
        for phone, name in contacts:
            if not phone:
                continue
            k = phone.replace("+", "").strip()
            nm = clean_name(name)
            if k in unique_contacts:
                if len(nm) > len(unique_contacts[k][1]):
                    unique_contacts[k] = (phone, nm)
            else:
                unique_contacts[k] = (phone, nm)
        return list(unique_contacts.values())
    except Exception as e:
        logger.error(f"Error in duplicate check: {str(e)}")
        return []

def classify_sms_delivery_status(status):
    canonical = SMS_STATUS_LOOKUP.get(str(status or "").strip().lower().replace(" ", ""))
    if canonical in SMS_PROMOTIONAL_STATUSES:
        return canonical, "promotional"
    if canonical in SMS_NON_PROMOTIONAL_STATUSES:
        return canonical, "non_promotional"
    if canonical in SMS_FAILED_STATUSES:
        return canonical, "failed"
    return str(status or "").strip(), "unrecognized"

def _find_sms_report_columns(df):
    normalized = {str(c).strip().lower(): c for c in df.columns}

    phone_col = None
    for key, col in normalized.items():
        if key in {"phone number", "phone", "msisdn", "mobile", "recipient"}:
            phone_col = col
            break
    if phone_col is None:
        for col in df.columns:
            sample = " ".join(df[col].dropna().astype(str).head(20).tolist())
            if re.search(r'\b(?:254|0)?[17]\d{8}\b', sample):
                phone_col = col
                break

    status_col = None
    for key, col in normalized.items():
        if key in {"delivery description", "status", "delivery status", "description"}:
            status_col = col
            break
    if status_col is None:
        for col in df.columns:
            values = df[col].dropna().astype(str).head(50).tolist()
            hits = sum(1 for value in values if classify_sms_delivery_status(value)[1] != "unrecognized")
            if hits:
                status_col = col
                break

    return phone_col, status_col

def parse_sms_delivery_report(file):
    name = (getattr(file, "name", "") or "").lower()
    try:
        if name.endswith(".csv"):
            frames = [pd.read_csv(file)]
        elif name.endswith((".xls", ".xlsx")):
            frames = list(_read_excel_all_sheets(file).values())
        else:
            raise ValueError("Unsupported SMS report format")
    except Exception as e:
        logger.error(f"SMS report read error: {str(e)}")
        st.error(f"Error reading SMS report: {str(e)}")
        return [], []

    rows = []
    issues = []
    seen = set()

    for df in frames:
        phone_col, status_col = _find_sms_report_columns(df)
        if phone_col is None or status_col is None:
            issues.append("Could not find phone/status columns in one sheet.")
            continue

        for _, row in df.iterrows():
            phone = format_phone_number(row.get(phone_col, ""))
            raw_status = row.get(status_col, "")
            canonical_status, category = classify_sms_delivery_status(raw_status)

            if not phone:
                continue

            dedupe_key = (phone, canonical_status)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            rows.append({
                "Phone": phone,
                "Raw Status": str(raw_status or "").strip(),
                "Delivery Status": canonical_status,
                "Category": category,
                "Promotional Allowed": category == "promotional",
                "Non Promotional Allowed": category in {"promotional", "non_promotional"},
                "Suppressed": category == "failed",
            })

    return rows, issues

def generate_standard_excel(data):
    """Generate standardized Excel output with openpyxl styling."""
    formatted_data = []
    seen = set()
    for item in data:
        if isinstance(item, dict):
            phone = item.get("Phone") or item.get("phone_number") or item.get("Phone or Email") or ""
            name = item.get("Name") or item.get("client_name") or ""
            phone = phone if str(phone).startswith("+") else format_phone_number(phone) or str(phone)
            if phone in seen:
                continue
            seen.add(phone)

            row = dict(item)
            clean_phone = phone.replace("+", "") if phone else ""
            if "Firstname(optional)" not in row or "Lastname(optional)" not in row:
                parts = str(name).strip().split() if name else []
                row.setdefault("Firstname(optional)", parts[0] if parts else "")
                row.setdefault("Lastname(optional)", " ".join(parts[1:]) if len(parts) > 1 else "")
            row.setdefault("Name", name)
            row.setdefault("Phone", phone)
            row.setdefault("Phone or Email", clean_phone)
            row.setdefault("phone(254)", clean_phone)
            row.setdefault("Valid", "Yes" if validate_contact(phone, str(name)) else "No")
            formatted_data.append(row)
            continue

        phone, name = item
        if phone in seen:
            continue
        seen.add(phone)
        first = last = ""
        if name:
            parts = name.strip().split()
            first = parts[0] if parts else ""
            last = " ".join(parts[1:]) if len(parts) > 1 else ""
        clean_phone = phone.replace("+", "") if phone else ""
        formatted_data.append({
            "Firstname(optional)": first,
            "Lastname(optional)": last,
            "Phone or Email": clean_phone,
            "phone(254)": clean_phone,
            "Valid": "Yes" if validate_contact(phone, name) else "No"
        })

    df = pd.DataFrame(formatted_data)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Contacts")
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        ws = writer.sheets["Contacts"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_alignment = Alignment(vertical="top", wrap_text=True)
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()])
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buffer.seek(0)
    return buffer, df
